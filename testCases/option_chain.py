import requests
import json
import time
from openpyxl import Workbook, load_workbook
import time
from datetime import date
import urllib.request

# -----------------------------------------------------------------------------
call_buy = 290
put_buy = 200

c_buy = 288.80
p_buy = 203.30

# -----------------------------------------------------------------------------


url = "https://www.nseindia.com/api/option-chain-indices?symbol=NIFTY"
headers = {"accept-encoding": "gzip, deflate, br",
           "accept - language": "en - US, en;q = 0.9",
           "referer": "https: // www.nseindia.com / option - chain",
           "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36"
           }
res = requests.get(url, headers=headers).text

data = json.loads(res)

exp_list = data['records']['expiryDates']

strick_price = data['records']['strikePrices']

strick = strick_price[47]

exp_date = exp_list[2]

for i in data['records']['data']:
    if i['expiryDate'] == exp_date and i['strikePrice'] == strick:
        ce = i['CE']['lastPrice']
        pe = i['PE']['lastPrice']

print(ce)
print(pe)

# ------------------------------------------------------------------------

file_path = 'TestData/Option_trade.xlsx'
# wbb = Workbook()
wb = load_workbook(file_path)
sheet = wb['Sheet']

print(sheet.max_row)
new_row = sheet.max_row + 1
# ce = call_ltp.split(".")
# pe = put_ltp.split(".")
# print(call_ltp[0:6])
profit = (call_buy - float(ce)) + (put_buy - float(pe))
total = profit * 50
print(total)

rahul_profit = (float(c_buy) - float(ce)) + (float(p_buy) - float(pe))
rahul_total = rahul_profit * 50

named_tuple = time.localtime()  # get struct_time
time_string = time.strftime("%H:%M", named_tuple)

# for i in data:
# sheet.append(sheet.max_row+1)

# sheet.cell(row=new_row, column=1).value = nifty_fut
sheet.cell(row=new_row, column=2).value = str(ce)
sheet.cell(row=new_row, column=3).value = str(pe)
sheet.cell(row=new_row, column=4).value = date.today()
sheet.cell(row=new_row, column=5).value = time_string
sheet.cell(row=new_row, column=6).value = round(total, 2)

wb.save('TestData/Option_trade.xlsx')
wb.close()

# ------------------------send to telegram----------------------------------------------
msg = "Chaitanya : Your profit is " + str(round(total, 2)) + " on time " + str(time_string) + " with CE at " + str(
    ce) + " and PE at " + str(pe)
url1 = 'https://api.telegram.org/bot6006884871:AAFqjs2rjTKfn7LYonjdmogq6v4-LAEegTU/sendMessage?chat_id=-894738745&text="{}"'.format(
    msg)
requests.get(url1)
webUrl = urllib.request.urlopen(url1)

msg1 = "Rahul : Your profit is " + str(round(rahul_total, 2)) + " on time " + str(time_string) + " with CE at " + str(
    ce) + " and PE at " + str(pe)
url2 = 'https://api.telegram.org/bot6006884871:AAFqjs2rjTKfn7LYonjdmogq6v4-LAEegTU/sendMessage?chat_id=-894738745&text="{}"'.format(
    msg1)
#         requests.get(url)
# driver.get(url2)
